from __future__ import annotations

import json
import os
import re
import threading
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Dict

from .base import ToolResult, ToolboxHelpers


# One process-wide lock for every mutating file operation. Multiple agents
# run concurrently against ONE shared workspace; without this, edit_file's
# read-modify-write silently loses a sibling agent's concurrent change
# (last-write-wins with both agents told "ok"). File I/O is fast relative to
# model/tool latency, so a single lock is contention-free in practice.
_FILE_MUTATION_LOCK = threading.RLock()


class FileTools(ToolboxHelpers):
    """Workspace file reading, writing, moving, and text inspection."""

    def _read_file(self, path: str, start_line: int = 0, line_count: int = 0) -> ToolResult:
        artifact_log = self._safe_artifact_log(path)
        if artifact_log:
            return ToolResult(True, "read_file", artifact_log.read_text(encoding="utf-8", errors="replace")[-12000:], {
                "path": str(artifact_log),
                "artifact_log": True,
                "relative_path": str(artifact_log.relative_to(self._artifacts_root().resolve())),
            })
        target = self._safe_path(path)
        if not target.exists() or not target.is_file():
            return ToolResult(False, "read_file", f"File not found: {path}", {})
        document = self._extract_document_text(target)
        if document is not None:
            text, kind = document
            return ToolResult(True, "read_file", text[:16000], {
                "path": str(target),
                "document_type": kind,
                "truncated": len(text) > 16000,
            })
        text = target.read_text(encoding="utf-8", errors="replace")
        if start_line > 0:
            # Windowed read with line numbers: lets the model navigate big
            # files deliberately instead of only ever seeing the tail.
            lines = text.splitlines()
            total = len(lines)
            begin = min(max(1, start_line), max(total, 1))
            count = line_count if line_count > 0 else 200
            window = lines[begin - 1:begin - 1 + count]
            numbered = "\n".join(f"{begin + idx}: {line}" for idx, line in enumerate(window))
            header = f"[{target.name}: lines {begin}-{begin + len(window) - 1} of {total}]"
            return ToolResult(True, "read_file", f"{header}\n{numbered}", {
                "path": str(target),
                "total_lines": total,
                "start_line": begin,
                "lines_shown": len(window),
            })
        if len(text) > 12000:
            total = len(text.splitlines())
            head = text[:2000]
            tail = text[-10000:]
            note = (
                f"[file is {len(text)} chars / {total} lines; showing first 2000 and last 10000 chars. "
                "Use read_file with start_line/line_count to read a specific region.]"
            )
            return ToolResult(True, "read_file", f"{note}\n{head}\n...\n{tail}", {
                "path": str(target),
                "truncated": True,
                "total_lines": total,
            })
        return ToolResult(True, "read_file", text, {"path": str(target)})

    def _validate_written_syntax(self, target: Path, content: str) -> str | None:
        """Deterministic guardrail: refuse to save a .py/.json file that no
        longer parses. Small models frequently produce broken edits and do
        not notice; a reject-with-reason turns silent corruption into an
        immediately recoverable step (SWE-agent/Aider lint-on-edit pattern)."""
        suffix = target.suffix.lower()
        try:
            if suffix == ".py":
                compile(content, str(target), "exec")
            elif suffix == ".json":
                # Many real .json files are actually JSONC (tsconfig.json,
                # .vscode/*.json, .eslintrc.json): comments and trailing commas
                # are valid to their tools but not to json.loads. Only strict-
                # validate when there are no comment markers, so we still catch
                # genuine typos in plain JSON without rejecting valid JSONC.
                if "//" not in content and "/*" not in content:
                    json.loads(content)
        except SyntaxError as exc:
            return f"line {exc.lineno}: {exc.msg}"
        except ValueError as exc:
            return str(exc)
        return None

    def _extract_document_text(self, target: Path) -> tuple[str, str] | None:
        """Extract readable text from binary documents (PDF/Word/Excel) so the
        model gets real content instead of garbled bytes. Returns None for
        plain-text files so they read normally."""
        suffix = target.suffix.lower()
        try:
            if suffix == ".pdf":
                from pypdf import PdfReader  # type: ignore

                reader = PdfReader(str(target))
                pages = [page.extract_text() or "" for page in reader.pages[:50]]
                return "\n\n".join(pages).strip(), "pdf"
            if suffix == ".docx":
                import docx  # type: ignore

                document = docx.Document(str(target))
                return "\n".join(p.text for p in document.paragraphs).strip(), "docx"
            if suffix in {".xlsx", ".xlsm"}:
                from openpyxl import load_workbook  # type: ignore

                workbook = load_workbook(str(target), read_only=True, data_only=True)
                lines: list[str] = []
                for sheet in workbook.worksheets:
                    lines.append(f"# Sheet: {sheet.title}")
                    for row in sheet.iter_rows(values_only=True):
                        lines.append("\t".join("" if cell is None else str(cell) for cell in row))
                return "\n".join(lines).strip(), "xlsx"
        except Exception as exc:
            return f"Could not extract {suffix} text: {exc}", suffix.lstrip(".") or "document"
        return None

    def _atomic_write_text(self, target: Path, content: str) -> None:
        """Write via temp file + os.replace so a concurrent reader can never
        observe a truncated half-written file. On Windows os.replace can fail
        with PermissionError when the destination is held open (a running
        server, an editor, AV); fall back to an in-place write and never leave
        a stray temp file behind."""
        temp = target.with_name(f".{target.name}.neo-tmp")
        try:
            temp.write_text(content, encoding="utf-8")
            os.replace(temp, target)
        except OSError:
            try:
                if temp.exists():
                    temp.unlink()
            except OSError:
                pass
            target.write_text(content, encoding="utf-8")

    def _write_file(self, path: str, content: str, force: bool = False) -> ToolResult:
        target = self._safe_path(path)
        scope_error = self._scope_violation(target)
        if scope_error:
            return ToolResult(False, "write_file", scope_error, {"path": str(target), "scope_blocked": True})
        if not force:
            syntax_error = self._validate_written_syntax(target, content)
            if syntax_error:
                return ToolResult(False, "write_file", (
                    f"Rejected: this content would break {target.suffix} syntax ({syntax_error}). "
                    "The file was NOT changed. Fix the content and retry, or pass force=true only if the file is intentionally invalid."
                ), {"path": str(target), "syntax_error": syntax_error})
        with _FILE_MUTATION_LOCK:
            before = target.read_text(encoding="utf-8", errors="replace") if target.exists() and target.is_file() else ""
            target.parent.mkdir(parents=True, exist_ok=True)
            self._atomic_write_text(target, content)
        meta = self._file_change_meta(target, before, content, "write_file")
        return ToolResult(True, "write_file", f"Wrote {len(content)} chars to {meta['relative_path']}", meta)

    def _append_file(self, path: str, content: str) -> ToolResult:
        target = self._safe_path(path)
        scope_error = self._scope_violation(target)
        if scope_error:
            return ToolResult(False, "append_file", scope_error, {"path": str(target), "scope_blocked": True})
        with _FILE_MUTATION_LOCK:
            before = target.read_text(encoding="utf-8", errors="replace") if target.exists() and target.is_file() else ""
            target.parent.mkdir(parents=True, exist_ok=True)
            with target.open("a", encoding="utf-8") as handle:
                handle.write(content)
        after = before + content
        meta = self._file_change_meta(target, before, after, "append_file")
        return ToolResult(True, "append_file", f"Appended {len(content)} chars to {meta['relative_path']}", meta)

    def _edit_file(self, path: str, old: str, new: str, replace_all: bool, force: bool = False) -> ToolResult:
        if not old:
            return ToolResult(False, "edit_file", "Old text is required.", {})
        target = self._safe_path(path)
        if not target.exists() or not target.is_file():
            return ToolResult(False, "edit_file", f"File not found: {path}", {})
        scope_error = self._scope_violation(target)
        if scope_error:
            return ToolResult(False, "edit_file", scope_error, {"path": str(target), "scope_blocked": True})
        with _FILE_MUTATION_LOCK:
            text = target.read_text(encoding="utf-8", errors="replace")
            count = text.count(old)
            if count == 0:
                return ToolResult(False, "edit_file", "Old text not found.", {"path": str(target)})
            if count > 1 and not replace_all:
                # Replacing "the first of N" silently edits a location the
                # model may not have meant. Make ambiguity a visible failure.
                return ToolResult(False, "edit_file", (
                    f"Old text matches {count} locations in {target.name}. "
                    "Include more surrounding context so it matches exactly once, or set replace_all=true."
                ), {"path": str(target), "matches": count})
            updated = text.replace(old, new) if replace_all else text.replace(old, new, 1)
            if not force:
                syntax_error = self._validate_written_syntax(target, updated)
                if syntax_error:
                    return ToolResult(False, "edit_file", (
                        f"Rejected: this edit would break {target.suffix} syntax ({syntax_error}). "
                        "The file was NOT changed. Fix the replacement and retry, or pass force=true only if the file is intentionally invalid."
                    ), {"path": str(target), "syntax_error": syntax_error})
            self._atomic_write_text(target, updated)
        changed = count if replace_all else 1
        meta = self._file_change_meta(target, text, updated, "edit_file")
        meta["replacements"] = changed
        return ToolResult(True, "edit_file", f"Replaced {changed} occurrence(s) in {meta['relative_path']}", meta)

    def _make_dir(self, path: str) -> ToolResult:
        target = self._safe_path(path)
        scope_error = self._scope_violation(target)
        if scope_error:
            return ToolResult(False, "make_dir", scope_error, {"path": str(target), "scope_blocked": True})
        target.mkdir(parents=True, exist_ok=True)
        relative = "." if target == self.workspace else str(target.relative_to(self.workspace))
        return ToolResult(True, "make_dir", f"Directory ready: {relative}", {
            "path": str(target),
            "relative_path": relative,
            "exists": target.exists(),
        })

    def _move_path(self, source: str, destination: str, overwrite: bool) -> ToolResult:
        import shutil

        src = self._safe_path(source)
        dst = self._safe_path(destination)
        scope_error = self._scope_violation(src) or self._scope_violation(dst)
        if scope_error:
            return ToolResult(False, "move_path", scope_error, {"source": str(src), "destination": str(dst), "scope_blocked": True})
        if not src.exists():
            return ToolResult(False, "move_path", f"Source not found: {source}", {"source": str(src)})
        if src == self.workspace:
            return ToolResult(False, "move_path", "Refusing to move the workspace root.", {"source": str(src)})
        if src == dst:
            return ToolResult(False, "move_path", "Source and destination are the same.", {"source": str(src), "destination": str(dst)})
        if src.is_dir() and (dst == src or src in dst.parents):
            return ToolResult(False, "move_path", "Refusing to move a directory inside itself.", {"source": str(src), "destination": str(dst)})

        final_target = dst / src.name if dst.exists() and dst.is_dir() else dst
        if final_target.exists():
            if not overwrite:
                return ToolResult(False, "move_path", f"Destination already exists: {final_target.relative_to(self.workspace)}", {
                    "source": str(src),
                    "destination": str(dst),
                    "final_path": str(final_target),
                })
            if final_target.is_dir():
                return ToolResult(False, "move_path", "Overwrite is only supported for files, not directories.", {
                    "source": str(src),
                    "destination": str(dst),
                    "final_path": str(final_target),
                })
            final_target.unlink()

        dst.parent.mkdir(parents=True, exist_ok=True)
        moved_to = Path(shutil.move(str(src), str(dst))).resolve()
        return ToolResult(True, "move_path", f"Moved {source} to {moved_to.relative_to(self.workspace)}", {
            "source": str(src),
            "destination": str(dst),
            "final_path": str(moved_to),
            "relative_path": str(moved_to.relative_to(self.workspace)),
            "moved": True,
        })

    def _copy_path(self, source: str, destination: str, recursive: bool) -> ToolResult:
        import shutil

        src = self._safe_path(source)
        dst = self._safe_path(destination)
        scope_error = self._scope_violation(dst)
        if scope_error:
            return ToolResult(False, "copy_path", scope_error, {"source": str(src), "destination": str(dst), "scope_blocked": True})
        if not src.exists():
            return ToolResult(False, "copy_path", f"Source not found: {source}", {"source": str(src)})
        if src == self.workspace:
            return ToolResult(False, "copy_path", "Refusing to copy the workspace root.", {"source": str(src)})
        if src.is_dir() and not recursive:
            return ToolResult(False, "copy_path", f"{source} is a directory. Pass recursive=true (cp -r) to copy it.", {
                "source": str(src),
                "is_dir": True,
            })

        final_target = dst / src.name if dst.exists() and dst.is_dir() else dst
        if final_target == src:
            return ToolResult(False, "copy_path", "Source and destination are the same.", {"source": str(src), "destination": str(dst)})
        if src.is_dir() and src in final_target.parents:
            return ToolResult(False, "copy_path", "Refusing to copy a directory inside itself.", {"source": str(src), "destination": str(final_target)})
        if final_target.exists():
            return ToolResult(False, "copy_path", f"Destination already exists: {final_target.relative_to(self.workspace)}", {
                "source": str(src),
                "destination": str(dst),
                "final_path": str(final_target),
            })

        final_target.parent.mkdir(parents=True, exist_ok=True)
        if src.is_dir():
            shutil.copytree(str(src), str(final_target))
        else:
            shutil.copy2(str(src), str(final_target))
        relative = str(final_target.relative_to(self.workspace))
        return ToolResult(True, "copy_path", f"Copied {source} to {relative}", {
            "source": str(src),
            "destination": str(dst),
            "final_path": str(final_target),
            "relative_path": relative,
            "copied": True,
            "is_dir": src.is_dir(),
        })

    def _delete_path(self, path: str, recursive: bool) -> ToolResult:
        import shutil

        target = self._safe_path(path)
        if target == self.workspace:
            return ToolResult(False, "delete_path", "Refusing to delete the workspace root.", {"path": str(target)})
        scope_error = self._scope_violation(target)
        if scope_error:
            return ToolResult(False, "delete_path", scope_error, {"path": str(target), "scope_blocked": True})
        if not target.exists():
            return ToolResult(False, "delete_path", f"Path not found: {path}", {"path": str(target)})
        relative = str(target.relative_to(self.workspace))
        if target.is_dir():
            has_children = any(target.iterdir())
            if has_children and not recursive:
                return ToolResult(
                    False,
                    "delete_path",
                    f"Directory {relative} is not empty. Pass recursive=true to delete it and its contents.",
                    {"path": str(target), "relative_path": relative, "is_dir": True},
                )
            shutil.rmtree(target) if has_children else target.rmdir()
        else:
            target.unlink()
        return ToolResult(True, "delete_path", f"Deleted {relative}", {
            "path": str(target),
            "relative_path": relative,
            "deleted": True,
        })

    def _file_change_meta(self, target: Path, before: str, after: str, operation: str) -> Dict[str, Any]:
        before_lines = before.splitlines()
        after_lines = after.splitlines()
        added = 0
        removed = 0
        for tag, i1, i2, j1, j2 in SequenceMatcher(None, before_lines, after_lines).get_opcodes():
            if tag == "insert":
                added += j2 - j1
            elif tag == "delete":
                removed += i2 - i1
            elif tag == "replace":
                removed += i2 - i1
                added += j2 - j1
        relative = str(target.relative_to(self.workspace))
        return {
            "path": str(target),
            "relative_path": relative,
            "file_name": target.name,
            "files_changed": 1,
            "added": added,
            "removed": removed,
            "operation": operation,
        }

    def _file_info(self, path: str) -> ToolResult:
        target = self._safe_path(path)
        if not target.exists():
            return ToolResult(False, "file_info", f"Path not found: {path}", {})
        stat = target.stat()
        payload = {
            "path": str(target.relative_to(self.workspace)) if target != self.workspace else ".",
            "type": "dir" if target.is_dir() else "file",
            "size": stat.st_size if target.is_file() else None,
            "modified": stat.st_mtime,
        }
        return ToolResult(True, "file_info", json.dumps(payload, indent=2), payload)

    def _tree(self, path: str, max_depth: int) -> ToolResult:
        root = self._safe_path(path)
        if not root.exists():
            return ToolResult(False, "tree", f"Path not found: {path}", {})
        max_depth = max(1, min(max_depth, 6))
        rows: list[str] = []

        def walk(current: Path, depth: int) -> None:
            if len(rows) >= 300:
                return
            rel = "." if current == self.workspace else str(current.relative_to(self.workspace))
            rows.append(("  " * depth) + (rel + ("/" if current.is_dir() else "")))
            if depth >= max_depth or not current.is_dir():
                return
            for child in sorted(current.iterdir(), key=lambda p: (not p.is_dir(), p.name.lower()))[:80]:
                walk(child, depth + 1)

        walk(root, 0)
        return ToolResult(True, "tree", "\n".join(rows), {"lines": len(rows), "max_depth": max_depth})

    def _list_files(self, path: str) -> ToolResult:
        target = self._safe_path(path)
        if not target.exists():
            return ToolResult(False, "list_files", f"Path not found: {path}", {})
        rows = []
        for item in sorted(target.iterdir(), key=lambda p: (not p.is_dir(), p.name.lower()))[:200]:
            rows.append({
                "name": item.name,
                "type": "dir" if item.is_dir() else "file",
                "size": item.stat().st_size if item.is_file() else None,
            })
        return ToolResult(True, "list_files", json.dumps(rows, indent=2), {"path": str(target), "count": len(rows)})

    def _search_files(self, path: str, pattern: str, max_results: int) -> ToolResult:
        target = self._safe_path(path)
        if not target.exists():
            return ToolResult(False, "search_files", f"Path not found: {path}", {})
        max_results = max(1, min(max_results, 500))
        matches = []
        for item in target.rglob(pattern or "*"):
            if len(matches) >= max_results:
                break
            if item.is_file():
                matches.append(str(item.relative_to(self.workspace)))
        return ToolResult(True, "search_files", json.dumps(matches, indent=2), {"count": len(matches), "pattern": pattern})

    def _grep(self, path: str, pattern: str, case_sensitive: bool, max_results: int) -> ToolResult:
        if not pattern:
            return ToolResult(False, "grep", "Pattern is required.", {})
        target = self._safe_path(path)
        if not target.exists():
            return ToolResult(False, "grep", f"Path not found: {path}", {})
        flags = 0 if case_sensitive else re.IGNORECASE
        regex = re.compile(pattern, flags)
        roots = [target] if target.is_file() else [p for p in target.rglob("*") if p.is_file()]
        max_results = max(1, min(max_results, 300))
        rows = []
        for file_path in roots:
            if len(rows) >= max_results:
                break
            if file_path.stat().st_size > 1_000_000:
                continue
            rel = str(file_path.relative_to(self.workspace))
            try:
                for line_no, line in enumerate(file_path.read_text(encoding="utf-8", errors="replace").splitlines(), 1):
                    if regex.search(line):
                        rows.append({"path": rel, "line": line_no, "text": line[:500]})
                        if len(rows) >= max_results:
                            break
            except OSError:
                continue
        return ToolResult(True, "grep", json.dumps(rows, indent=2), {"count": len(rows), "pattern": pattern})

    def _python_symbols(self, path: str) -> ToolResult:
        import ast

        target = self._safe_path(path)
        if not target.exists() or not target.is_file():
            return ToolResult(False, "python_symbols", f"File not found: {path}", {})
        tree = ast.parse(target.read_text(encoding="utf-8", errors="replace"))
        rows = []
        for node in ast.walk(tree):
            if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef, ast.ClassDef)):
                rows.append({"kind": type(node).__name__, "name": node.name, "line": node.lineno})
            elif isinstance(node, ast.Import):
                rows.extend({"kind": "Import", "name": alias.name, "line": node.lineno} for alias in node.names)
            elif isinstance(node, ast.ImportFrom):
                module = node.module or ""
                rows.extend({"kind": "ImportFrom", "name": f"{module}.{alias.name}", "line": node.lineno} for alias in node.names)
        rows = sorted(rows, key=lambda item: (item["line"], item["kind"], item["name"]))[:300]
        return ToolResult(True, "python_symbols", json.dumps(rows, indent=2), {"count": len(rows)})

    def _json_validate(self, text: str) -> ToolResult:
        try:
            parsed = json.loads(text)
        except json.JSONDecodeError as exc:
            return ToolResult(False, "json_validate", f"Invalid JSON at line {exc.lineno}, column {exc.colno}: {exc.msg}", {})
        summary = {
            "valid": True,
            "type": type(parsed).__name__,
            "keys": list(parsed.keys())[:50] if isinstance(parsed, dict) else None,
            "items": len(parsed) if isinstance(parsed, (dict, list)) else None,
        }
        return ToolResult(True, "json_validate", json.dumps(summary, indent=2), summary)

    def _safe_artifact_log(self, path: str) -> Path | None:
        if not path:
            return None
        candidate = Path(path).resolve()
        root = self._artifacts_root().resolve()
        if candidate != root and root not in candidate.parents:
            return None
        if candidate.suffix != ".log" or not candidate.exists():
            return None
        return candidate
